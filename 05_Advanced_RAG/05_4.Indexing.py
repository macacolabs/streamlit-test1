"""
RAG 인덱싱 파이프라인 — Contextual Retrieval 고도화

적용 기법:
  1. 문서 종류별 인덱싱 전략 분리
  2. 청킹 전 구조 보존 전처리
  3. 확장 메타데이터 + 증분 인덱싱
  4. [NEW] Contextual Embedding
       - 각 청크에 LLM으로 맥락 설명을 생성해 page_content 앞에 prepend
       - 임베딩이 청크 고유 내용 + 문서 내 위치를 모두 포착
  5. [NEW] Contextual BM25
       - 맥락 보강된 텍스트로 BM25 역인덱스 구축 (bm25_index.pkl)
       - 05_5.Retrieval의 하이브리드 검색(Semantic + BM25 + RRF)에 사용
"""

import hashlib
import importlib.util
import os
import pickle
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from textwrap import dedent

from langchain_chroma import Chroma
from langchain_community.document_loaders import PyPDFLoader
from langchain_core.documents import Document
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI, OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from rank_bm25 import BM25Okapi

# ──────────────────────────────────────────────
# 형제 모듈 동적 임포트
# ──────────────────────────────────────────────

def _load_sibling(filename: str):
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    spec = importlib.util.spec_from_file_location(filename, path)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_auth    = _load_sibling("05_3.Auth.py")
_schemas = _load_sibling("05_2.Schemas.py")

DATA_DIR        = _auth.DATA_DIR
PDF_PATH        = _auth.PDF_PATH
VECTOR_DB_PATH  = _auth.VECTOR_DB_PATH
BM25_INDEX_PATH = _auth.BM25_INDEX_PATH
COLLECTION_NAME = _auth.COLLECTION_NAME
TYPE_INFO       = _schemas.TYPE_INFO
TYPE_MARKERS    = _schemas.TYPE_MARKERS


# ──────────────────────────────────────────────
# 공통 유틸리티
# ──────────────────────────────────────────────

def _file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(65536), b""):
            h.update(block)
    return h.hexdigest()[:16]


def _tokenize(text: str) -> list[str]:
    """BM25용 한국어/영어 토크나이저 (형태소 분석 없이 어절 기반)."""
    return re.findall(r"[가-힣]{2,}|[A-Za-z]{2,}|\d+", text.lower())


_STOPWORDS = {
    "이", "그", "저", "것", "수", "등", "및", "을", "를", "가", "은",
    "는", "의", "에", "로", "으로", "하다", "있다", "되다", "하는", "있는",
    "되는", "한", "그리고", "또한", "통해", "위해", "대한", "따라", "기반",
    "위한", "통한", "대해", "관련", "경우", "때문", "이후", "이전", "이상",
}


def _extract_keywords(text: str, n: int = 6) -> str:
    tokens = re.findall(r"[가-힣]{2,}", text)
    freq: dict[str, int] = {}
    for t in tokens:
        if t not in _STOPWORDS:
            freq[t] = freq.get(t, 0) + 1
    top = sorted(freq, key=lambda x: freq[x], reverse=True)[:n]
    return ",".join(top)


def _clean_pdf_text(text: str) -> str:
    text = re.sub(r"(\w)-\n(\w)", r"\1\2", text)
    text = re.sub(r"(?m)^\s*\d{1,4}\s*$", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return "\n".join(ln.rstrip() for ln in text.split("\n")).strip()


def _tag_chunks(chunks: list[Document], source_file: str,
                file_hash: str, base_meta: dict) -> list[Document]:
    now   = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    total = len(chunks)
    for i, chunk in enumerate(chunks):
        kw = _extract_keywords(chunk.page_content)
        chunk.metadata.update({
            **base_meta,
            "source_file":  source_file,
            "file_hash":    file_hash,
            "chunk_index":  i,
            "total_chunks": total,
            "char_count":   len(chunk.page_content),
            "keywords":     kw,
            "indexed_at":   now,
        })
    return chunks


# ──────────────────────────────────────────────
# Contextual Embedding (기법 4)
# ──────────────────────────────────────────────

_CTX_SYSTEM = (
    "당신은 RAG 검색 최적화 전문가다. "
    "지시한 형식의 맥락 설명만 출력하고 다른 말은 붙이지 마라."
)

_CTX_PROMPT = dedent("""
    아래는 원본 문서(또는 섹션)의 내용이다:
    <document>
    {doc_text}
    </document>

    아래는 위 문서에서 추출한 청크다:
    <chunk>
    {chunk_text}
    </chunk>

    이 청크가 문서 전체 맥락에서 어떤 내용인지 2~3문장으로 설명하라.
    포함 사항:
    - 어떤 문서/섹션에 속하는지
    - 어떤 개념·주제와 연관되는지
    - 검색 시 유용한 핵심 용어
    설명만 출력하라.
""").strip()


def _generate_context(llm: ChatOpenAI, doc_text: str, chunk_text: str) -> str:
    """청크 1개에 대한 맥락 설명을 LLM으로 생성한다."""
    prompt = _CTX_PROMPT.format(
        doc_text=doc_text[:2500],
        chunk_text=chunk_text[:700],
    )
    try:
        resp = llm.invoke([
            SystemMessage(content=_CTX_SYSTEM),
            HumanMessage(content=prompt),
        ])
        return resp.content.strip()
    except Exception as e:
        print(f"[Contextual] 컨텍스트 생성 실패: {e}")
        return ""


def _apply_contextual_embedding(
    chunks: list[Document],
    api_key: str,
    max_workers: int = 8,
) -> list[Document]:
    """
    청크 리스트 전체에 Contextual Embedding을 적용한다.

    각 청크 metadata["_doc_context"] 를 읽어 맥락을 생성한 뒤:
      - page_content  = "<맥락>\n\n<원본 청크>"  (임베딩·BM25 검색용)
      - original_content = <원본 청크>           (표시용)
      - has_context   = True / False
      - _doc_context 키 제거 (ChromaDB에 저장하지 않음)
    """
    llm = ChatOpenAI(model="gpt-4.1-mini", temperature=0, api_key=api_key)
    print(f"[Contextual] {len(chunks)}개 청크 맥락 생성 시작 (병렬 {max_workers}개)...")

    def _process(item: tuple[int, Document]) -> tuple[int, Document]:
        idx, chunk = item
        doc_ctx = chunk.metadata.pop("_doc_context", "")
        original = chunk.page_content
        ctx = _generate_context(llm, doc_ctx, original) if doc_ctx else ""
        chunk.page_content = f"{ctx}\n\n{original}" if ctx else original
        chunk.metadata["original_content"] = original
        chunk.metadata["has_context"]      = bool(ctx)
        return idx, chunk

    results: list[Document | None] = [None] * len(chunks)
    done_count = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_process, (i, c)): i
                   for i, c in enumerate(chunks)}
        for future in as_completed(futures):
            idx, chunk = future.result()
            results[idx] = chunk
            done_count += 1
            if done_count % 20 == 0 or done_count == len(chunks):
                print(f"[Contextual] {done_count}/{len(chunks)} 완료")

    # 타입 좁히기 (None이 남아 있으면 원본으로 폴백)
    return [r if r is not None else chunks[i] for i, r in enumerate(results)]


# ──────────────────────────────────────────────
# Contextual BM25 (기법 5) — 인덱스 관리
# ──────────────────────────────────────────────

def _rebuild_bm25(vectorstore: Chroma, path: str) -> None:
    """
    ChromaDB에 현재 저장된 모든 청크로 BM25 인덱스를 (재)구축하고 저장한다.
    ChromaDB가 단일 진실 소스(source of truth)이므로 항상 일관성이 보장된다.
    """
    print("[BM25] 인덱스 재구축 중...")
    result    = vectorstore._collection.get(include=["documents", "metadatas"])
    contents  = result["documents"]
    metadatas = result["metadatas"]

    if not contents:
        print("[BM25] 저장할 문서가 없음. 스킵.")
        return

    tokenized = [_tokenize(text) for text in contents]
    bm25_obj  = BM25Okapi(tokenized)

    data = {"bm25": bm25_obj, "contents": contents, "metadatas": metadatas}
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        pickle.dump(data, f)

    print(f"[BM25] 인덱스 저장 완료 ({len(contents)}개 항목) → {path}")


def load_bm25(path: str) -> dict | None:
    """BM25 인덱스를 파일에서 로드한다. 파일이 없으면 None 반환."""
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        data = pickle.load(f)
    print(f"[BM25] 인덱스 로드 완료 ({len(data['contents'])}개 항목)")
    return data


# ──────────────────────────────────────────────
# 전략 1: AXCompass.pdf — 유형 섹션 경계 기반 분리
# ──────────────────────────────────────────────

_AX_SPLITTER = RecursiveCharacterTextSplitter(
    chunk_size=600,
    chunk_overlap=80,
    separators=["\n\n", "\n", "。", ".", " "],
)


def _load_ax_compass(pdf_path: str, fhash: str) -> list[Document]:
    pages     = PyPDFLoader(pdf_path).load()
    full_text = _clean_pdf_text("\n".join(p.page_content for p in pages))

    marker_positions: list[tuple[int, str]] = []
    for type_name, marker in TYPE_MARKERS.items():
        idx = full_text.find(marker)
        if idx >= 0:
            marker_positions.append((idx, type_name))
    marker_positions.sort(key=lambda x: x[0])

    if not marker_positions:
        print("[RAG][AX] 섹션 마커를 찾지 못했습니다. 기본 청킹으로 전환.")
        chunks = _AX_SPLITTER.split_documents(pages)
        for c in chunks:
            c.metadata.update({"doc_type": "ax_compass", "_doc_context": full_text[:2500]})
        return _tag_chunks(chunks, "AXCompass.pdf", fhash, {"doc_type": "ax_compass"})

    all_chunks: list[Document] = []
    for i, (start, type_name) in enumerate(marker_positions):
        end          = marker_positions[i + 1][0] if i + 1 < len(marker_positions) else len(full_text)
        section_text = full_text[start:end].strip()
        if not section_text:
            continue

        info    = TYPE_INFO.get(type_name, {})
        sec_doc = Document(
            page_content=section_text,
            metadata={
                "doc_type":    "ax_compass",
                "type_name":   type_name,
                "group":       info.get("group", ""),
                "english":     info.get("english", ""),
                "section":     type_name,
                "_doc_context": section_text[:2500],  # 섹션 자체가 맥락 소스
            },
        )
        chunks = _AX_SPLITTER.split_documents([sec_doc])
        all_chunks.extend(chunks)

    fname = os.path.basename(pdf_path)
    print(f"[RAG][AX] {fname}: {len(marker_positions)}개 섹션 → {len(all_chunks)}개 청크")
    return _tag_chunks(all_chunks, fname, fhash, {})


# ──────────────────────────────────────────────
# 전략 2: 커리큘럼 PDF — Day/모듈 헤딩 구조 보존
# ──────────────────────────────────────────────

_CURRICULUM_SPLITTER = RecursiveCharacterTextSplitter(
    chunk_size=800,
    chunk_overlap=150,
    separators=["\n\n", "\n", "。", ".", " "],
)

_SECTION_PATTERN = re.compile(
    r"(?m)^(?:"
    r"Day\s*\d+|일차\s*\d+|"
    r"모듈\s*\d+|Module\s*\d+|"
    r"Chapter\s*\d+|챕터\s*\d+|"
    r"\d+\.\s+[^\n]{5,}"
    r")",
    re.IGNORECASE,
)


def _estimate_page(full_text: str, char_pos: int, page_texts: list[str]) -> int:
    cumulative = 0
    for i, pt in enumerate(page_texts):
        cumulative += len(pt) + 2
        if char_pos < cumulative:
            return i + 1
    return len(page_texts)


def _load_curriculum_pdf(pdf_path: str, fhash: str, course_name: str) -> list[Document]:
    pages      = PyPDFLoader(pdf_path).load()
    page_texts = [_clean_pdf_text(p.page_content) for p in pages]
    full_text  = "\n\n".join(page_texts)

    # 전체 문서 요약 (맥락 생성 시 문서 식별에 사용)
    doc_summary = f"[커리큘럼: {course_name}]\n{full_text[:1500]}"

    matches   = list(_SECTION_PATTERN.finditer(full_text))
    base_meta = {"doc_type": "curriculum_example", "course_name": course_name}
    all_chunks: list[Document] = []

    if not matches:
        for p_idx, page in enumerate(pages):
            page.metadata.update({**base_meta, "page_start": p_idx + 1, "section": "",
                                   "_doc_context": doc_summary})
            chunks = _CURRICULUM_SPLITTER.split_documents([page])
            all_chunks.extend(chunks)
    else:
        boundaries = [(m.start(), m.group().strip()) for m in matches]
        for i, (start, heading) in enumerate(boundaries):
            end          = boundaries[i + 1][0] if i + 1 < len(boundaries) else len(full_text)
            section_text = full_text[start:end].strip()
            if not section_text:
                continue
            page_num = _estimate_page(full_text, start, page_texts)
            # 맥락 = 문서 요약 + 섹션 제목 + 섹션 내용 앞부분
            doc_ctx  = f"{doc_summary}\n\n[섹션: {heading}]\n{section_text[:1000]}"
            sec_doc  = Document(
                page_content=section_text,
                metadata={**base_meta, "section": heading, "page_start": page_num,
                           "_doc_context": doc_ctx[:2500]},
            )
            chunks = _CURRICULUM_SPLITTER.split_documents([sec_doc])
            all_chunks.extend(chunks)

    fname = os.path.basename(pdf_path)
    print(f"[RAG][PDF] {fname}: {len(matches) or len(pages)}개 섹션 → {len(all_chunks)}개 청크")
    return _tag_chunks(all_chunks, fname, fhash, {})


# ──────────────────────────────────────────────
# 전략 3: 커리큘럼 Excel — 헤더+행 쌍 재구성
# ──────────────────────────────────────────────

_EXCEL_SPLITTER = RecursiveCharacterTextSplitter(
    chunk_size=600,
    chunk_overlap=60,
    separators=["\n\n", "\n", " "],
)


def _load_curriculum_excel(xlsx_path: str, fhash: str, course_name: str) -> list[Document]:
    try:
        import openpyxl
    except ImportError:
        from langchain_community.document_loaders import UnstructuredExcelLoader
        docs  = UnstructuredExcelLoader(xlsx_path).load()
        fname = os.path.basename(xlsx_path)
        base  = {"doc_type": "curriculum_example", "course_name": course_name,
                 "section": "", "page_start": 1}
        for doc in docs:
            doc.metadata.update(base)
            doc.metadata["_doc_context"] = f"[커리큘럼: {course_name}]\n{doc.page_content[:2000]}"
        chunks = _EXCEL_SPLITTER.split_documents(docs)
        print(f"[RAG][XLS] {fname}: fallback → {len(chunks)}개 청크")
        return _tag_chunks(chunks, fname, fhash, {})

    wb         = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    fname      = os.path.basename(xlsx_path)
    base       = {"doc_type": "curriculum_example", "course_name": course_name}
    all_chunks : list[Document] = []
    BUCKET_LIMIT = 1200

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers   = [str(h).strip() if h is not None else f"열{i+1}"
                     for i, h in enumerate(rows[0])]
        row_texts : list[str] = []
        for row in rows[1:]:
            pairs = [f"{h}: {str(v).strip()}"
                     for h, v in zip(headers, row)
                     if v is not None and str(v).strip()]
            if pairs:
                row_texts.append("\n".join(pairs))

        if not row_texts:
            continue

        # 전체 시트 텍스트를 맥락 소스로 사용
        full_sheet = "\n\n".join(row_texts)
        doc_ctx    = f"[커리큘럼: {course_name} / 시트: {sheet_name}]\n{full_sheet[:2000]}"

        bucket: list[str] = []
        bucket_len = 0
        for rt in row_texts:
            if bucket and bucket_len + len(rt) > BUCKET_LIMIT:
                all_chunks.append(Document(
                    page_content="\n\n".join(bucket),
                    metadata={**base, "section": sheet_name, "page_start": 1,
                               "_doc_context": doc_ctx},
                ))
                bucket, bucket_len = [], 0
            bucket.append(rt)
            bucket_len += len(rt)
        if bucket:
            all_chunks.append(Document(
                page_content="\n\n".join(bucket),
                metadata={**base, "section": sheet_name, "page_start": 1,
                           "_doc_context": doc_ctx},
            ))

    wb.close()
    print(f"[RAG][XLS] {fname}: {len(wb.sheetnames)}개 시트 → {len(all_chunks)}개 청크")
    return _tag_chunks(all_chunks, fname, fhash, {})


# ──────────────────────────────────────────────
# 문서 수집 (파일별 hash + chunks)
# ──────────────────────────────────────────────

def _collect_documents() -> dict[str, tuple[str, list[Document]]]:
    result: dict[str, tuple[str, list[Document]]] = {}

    if not os.path.exists(PDF_PATH):
        raise FileNotFoundError(f"AXCompass.pdf 파일을 찾을 수 없습니다: {PDF_PATH}")

    fhash = _file_hash(PDF_PATH)
    fname = os.path.basename(PDF_PATH)
    result[fname] = (fhash, _load_ax_compass(PDF_PATH, fhash))

    for fn in sorted(os.listdir(DATA_DIR)):
        if fn.endswith(".pdf") and fn != "AXCompass.pdf":
            fpath = os.path.join(DATA_DIR, fn)
            fhash = _file_hash(fpath)
            result[fn] = (fhash, _load_curriculum_pdf(fpath, fhash, os.path.splitext(fn)[0]))

    for fn in sorted(os.listdir(DATA_DIR)):
        if fn.endswith(".xlsx"):
            fpath = os.path.join(DATA_DIR, fn)
            fhash = _file_hash(fpath)
            result[fn] = (fhash, _load_curriculum_excel(fpath, fhash, os.path.splitext(fn)[0]))

    return result


# ──────────────────────────────────────────────
# 증분 인덱싱 헬퍼
# ──────────────────────────────────────────────

def _get_indexed_hashes(vectorstore: Chroma) -> dict[str, str]:
    if vectorstore._collection.count() == 0:
        return {}
    results = vectorstore._collection.get(include=["metadatas"])
    hashes: dict[str, str] = {}
    for meta in results["metadatas"]:
        fn, fh = meta.get("source_file"), meta.get("file_hash")
        if fn and fh:
            hashes[fn] = fh
    return hashes


def _delete_file_chunks(vectorstore: Chroma, source_file: str) -> None:
    vectorstore._collection.delete(where={"source_file": {"$eq": source_file}})


# ──────────────────────────────────────────────
# RAG 2단계: 벡터 DB 구축 (증분 + Contextual)
# ──────────────────────────────────────────────

def setup_vector_store(
    api_key: str,
    force_reindex: bool = False,
    enable_contextual: bool = True,
) -> Chroma:
    """
    벡터 DB와 BM25 인덱스를 구축/증분 업데이트한다.

    Parameters
    ----------
    api_key            : OpenAI API 키
    force_reindex      : True이면 기존 컬렉션 전체 삭제 후 재구축
    enable_contextual  : True이면 Contextual Embedding 적용 (LLM 호출 발생)
    """
    embeddings = OpenAIEmbeddings(model="text-embedding-3-small", api_key=api_key)
    vectorstore = Chroma(
        collection_name=COLLECTION_NAME,
        embedding_function=embeddings,
        persist_directory=VECTOR_DB_PATH,
    )

    if force_reindex and vectorstore._collection.count() > 0:
        print("[VectorDB] 강제 재인덱싱: 기존 컬렉션 전체 삭제...")
        vectorstore._collection.delete(where={})
        # BM25도 초기화
        if os.path.exists(BM25_INDEX_PATH):
            os.remove(BM25_INDEX_PATH)

    indexed_hashes  = _get_indexed_hashes(vectorstore)
    file_docs       = _collect_documents()
    changed_files   = 0

    for fname, (fhash, chunks) in file_docs.items():
        existing_hash = indexed_hashes.get(fname)

        if existing_hash == fhash:
            print(f"[VectorDB] 스킵 (변경 없음): {fname}")
            continue

        if existing_hash is not None:
            print(f"[VectorDB] 업데이트: {fname}")
            _delete_file_chunks(vectorstore, fname)
        else:
            print(f"[VectorDB] 신규 추가: {fname}")

        # Contextual Embedding 적용 (파일 단위)
        if enable_contextual and chunks:
            chunks = _apply_contextual_embedding(chunks, api_key)
        else:
            # _doc_context 메타데이터 제거 (ChromaDB에 저장하지 않음)
            for c in chunks:
                c.metadata.pop("_doc_context", None)
                c.metadata["has_context"] = False
                c.metadata["original_content"] = c.page_content

        if chunks:
            vectorstore.add_documents(chunks)
        changed_files += 1

    total = vectorstore._collection.count()
    print(f"[VectorDB] 완료 — 변경: {changed_files}개 파일 | 총 {total}개 청크")

    # BM25: 변경이 있으면 전체 재구축 (ChromaDB가 단일 진실 소스)
    if changed_files > 0 or not os.path.exists(BM25_INDEX_PATH):
        _rebuild_bm25(vectorstore, BM25_INDEX_PATH)

    return vectorstore
